import sys
import pandas as pd
import json
import PyPDF2
import os
import math
from openpyxl import Workbook


wb = Workbook()
ws = wb.active
arg = sys.argv[1]
#pname = os.path.join(os.getcwd(),arg+'.pdf')
#pdfobj = PyPDF2.PdfFileReader(pname)
# arg = sys.argv[1]
#fname = "D:\\Python\\untitled1\\"
# fname = os.getcwd()+'\\'+arg + ".json"
print(arg)
fname = arg

with open(fname, 'r') as f:
  distros_dict = json.load(f)

d = []
k = []
a = 1

# print (type(distros_dict))
if isinstance(distros_dict, dict):
  for blocks in distros_dict['Blocks']:
    if blocks['BlockType'] == 'WORD':
      d = [blocks['Page'], blocks['Text'], blocks['Geometry']['BoundingBox']['Top'],
           blocks['Geometry']['BoundingBox']['Left'], blocks['Geometry']['BoundingBox']['Width'],
           blocks['Geometry']['BoundingBox']['Height']]
      k.append(d)
else:
  for lis in distros_dict:
    for blocks in lis['Blocks']:
      if blocks['BlockType'] == 'LINE':
        d = [blocks['Page'], blocks['Text'], blocks['Geometry']['BoundingBox']['Top'],
             blocks['Geometry']['BoundingBox']['Left'], blocks['Geometry']['BoundingBox']['Width'],
             blocks['Geometry']['BoundingBox']['Height']]
        k.append(d)

dk = pd.DataFrame(k)
dk.columns = ['Page','Text','Top','Left','width','height']
dk["line"] = int(0)

lin = 1
for row in dk.index:
  if row > 0:
    if (dk['Page'][row] > dk['Page'][row-1]):
      lin = 1
    #if (dk['Page'][row] == 3):
    #print(dk['Top'][row-1],dk['Top'][row],(dk['Top'][row] - dk['Top'][row - 1]),dk['Text'][row - 1])
    #if (dk['Top'][row] > dk['Top'][row-1]) >= 0.008 :#or (dk['Left'][row-1] - dk['Left'][row] > 0.50) :
    if ((dk['Left'][row] < dk['Left'][row - 1]) and (dk['Page'][row] == dk['Page'][row - 1])):
      lin += 1
    #dk["line"][row] = lin
    dk._set_value(row,["line"],lin)
  else:
    dk._set_value(row, ["line"], lin)
    lin += 1
    #dk["line"][row] = lin

dk['col'] = int(0)
col = 0
inc = 0
for row in dk.index:
  if row > 0:
    if (dk["line"][row] > dk["line"][row-1]) or (dk['Page'][row] > dk['Page'][row-1]):
      col = 1
      inc += 1
      dk._set_value(inc, ["col"], col)
    else:
      inc += 1
      col += 1
      dk._set_value(inc, ["col"], col)
  else:
    col = 1
    dk._set_value(inc, ["col"], col)

#dk.to_csv("mcare_df.csv")
dk['pos'] = int(0)
ws = wb.create_sheet("Page1")
py = 0
ln = 0
pstr1 = ''

for row in dk.index:
  #if dk['Page'][]
  x = int(dk['line'][row])
  y = int(dk['pos'][row])
  z = int(dk['Page'][row])

  str1 = dk.query('line==@x and pos==@y and Page==@z')['Text'].iloc[0]
  ny = int(dk['Left'][row] * 15)
  ny = math.trunc(ny)


  if ny <=0:
    ny = 1
  else:
    ny +=1
    pstr1 = str1

  if py == ny:
   #ln = dk['line'][row]
   #print(pstr1,str1)
   ny += 1
  else:
    py = ny
    if dk['line'][row] == ln:
      ny += 1

  dk._set_value(row,['pos'],ny)

  if row > 0 and dk['Page'][row] > dk['Page'][row-1]:
    ws = wb.create_sheet("Page"+str(dk["Page"][row]))
  ws.cell(row=x, column=ny).value = str1

wb.remove(wb["Sheet"])
out_file = '{}.xlsx'.format(os.path.splitext(os.path.basename(arg))[0])

wb.save(os.path.join(os.getcwd(),'media','xlsx',out_file))





